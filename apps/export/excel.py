import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from apps.catalog.models import Product


def generate_price_excel(user):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Прайс-лист SplitHome'

    headers = ['Артикул', 'НС-код', 'Название', 'Бренд', 'Категория',
               'Остаток', 'РИЦ', 'Опт. цена']
    header_fill = PatternFill(start_color='2E7CF6', end_color='2E7CF6', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    products = Product.objects.filter(is_active=True).select_related(
        'brand', 'category', 'stock'
    ).order_by('brand__title', 'title')

    for row, p in enumerate(products, 2):
        stock_qty = p.stock.quantity if hasattr(p, 'stock') else 0
        wholesale = user.get_wholesale_price(p.price_wholesale) if p.price_wholesale else ''
        ws.append([
            p.articul, p.nc_code, p.title,
            str(p.brand) if p.brand else '',
            str(p.category) if p.category else '',
            stock_qty,
            float(p.ric) if p.ric else '',
            float(wholesale) if wholesale else '',
        ])

    col_widths = [15, 15, 50, 20, 25, 10, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
